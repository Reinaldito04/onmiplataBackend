import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
import os

class PagosExcel:
    def __init__(self, filename):
        self.filename = filename
        self.data = []
        self.year = ''

    def write_to_excel(self, output_path):
        # Crear un nuevo libro de trabajo
        wb = Workbook()
        ws = wb.active
        ws.title = "Pagos"

        # Establecer los encabezados de las columnas en español
        headers = [
            "Nombre", "Apellido", "DNI", "Monto", "Fecha", "ID Contrato", "ID", "Tipo de Pago", "Método de Pago"
        ]
        
        row_idx = 1
        ws.cell(row=row_idx,column=5,value=f"Pagos del {self.year}")
        row_idx +=2
        
        # Estilo de encabezado
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='33d1ff', end_color='33d1ff', fill_type='solid')  # Amarillo
        header_border = Border(bottom=Side(style='thin'))
    
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border

        # Estilo de datos
        data_font = Font()
        data_border = Border(bottom=Side(style='thin'))

        # Agregar los datos
        for row in self.data:
            ws.append([
                row['Name'], row['Lastname'], row['DNI'], row['Amount'], 
                row['Date'], row['IdContract'], 
                row['ID'], row['TypePay'], row['PaymentMethod']
            ])

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.font = data_font
                cell.border = data_border

        # Ajustar ancho de las columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Agregar imagen si existe
        
        if not os.path.exists(output_path):
            os.makedirs(output_path)
        # Guardar el archivo
        nombreArchivo = f'Pagos - {self.year}.xlsx'
        rutaCompleta = os.path.join(output_path,nombreArchivo)
        
        wb.save(rutaCompleta)
        return {"nombre_archivo": nombreArchivo, "output_path": rutaCompleta}
    def set_data(self, data):
        self.data = data
    
    def set_year(self,year):
        self.year = year
    

# Ejemplo de uso
if __name__ == "__main__":
    # Suponiendo que los datos vienen de alguna fuente
    example_data = [
        {"Name": "John", "Lastname": "Doe", "DNI": "123456789", "Amount": 1000, "Date": "01/01/2024", "IdContract": 1, "PaymentType": "Empresa", "ID": 1, "TypePay": "Alquiler", "PaymentMethod": "Transferencia"},
        {"Name": "Jane", "Lastname": "Smith", "DNI": "987654321", "Amount": 1500, "Date": "01/02/2024", "IdContract": 2, "PaymentType": "Personal", "ID": 2, "TypePay": "Honorarios", "PaymentMethod": "Efectivo"}
    ]

    pagos_excel = PagosExcel(filename="Pagos.xlsx")
    pagos_excel.set_data(example_data)
    pagos_excel.set_year('2024')
    pagos_excel.write_to_excel("Pagos.xlsx")
