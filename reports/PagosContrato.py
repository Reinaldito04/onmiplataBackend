from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, numbers
from openpyxl.drawing.image import Image
import datetime
import os
class ContratoInquilinoExcel:
    def __init__(self, filename):
        self.filename = filename
        self.imagen = 'media/assets/logo.png'
        self.datos_inquilino = {
            "INQUILINO": "",
            "N° CONTACTO": "",
            "CORREO": "",
            "INMUEBLE": "",
            "FECHA DE CONTRATO": "",
            "CANON": "",
            "DEPOSITO EN GARANTIA": ""
        }
        self.canones_mensuales = []
        self.total_contrato = ""
        self.depositos_efectuados = []
        self.total_depositado = ""
      
    
    def set_datos_inquilino(self, datos):
        self.datos_inquilino.update(datos)
    
    def set_canones_mensuales(self, canones):
        self.canones_mensuales = canones
    
    def set_total_contrato(self, total):
        self.total_contrato = total
    
    def set_depositos_efectuados(self, depositos):
        self.depositos_efectuados = depositos
    
    def set_total_depositado(self, total):
        self.total_depositado = total
    
    def write_to_excel(self,output_path):
        # Crear un nuevo libro de trabajo
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Contrato Inquilino'
       
        # Estilos
        header_style = Font(bold=True)
        border_top = Border(top=Side(style='thin'))
        cell_alignment = Alignment(horizontal='center', vertical='center')
        currency_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        inquilino = Font(color='ff0000')
        # Escribir datos en el archivo Excel
        underline_style = Font(underline="double")
        row_idx = 1
        
        logo = Image(self.imagen)
        worksheet.add_image(logo, 'A1')
        row_idx +=6
         # Escribir datos del inquilino con subrayado doble en el nombre del campo
        for key, value in self.datos_inquilino.items():
            if key == "INQUILINO":
                worksheet.cell(row=row_idx, column=1, value="INQUILINO").font = header_style
                worksheet.cell(row=row_idx, column=2, value=value).font = inquilino
            if key == 'N° CONTACTO':
                worksheet.cell(row=row_idx, column=1, value="N° CONTACTO").font = header_style
                worksheet.cell(row=row_idx, column=2, value=value).font = inquilino
            else:
                worksheet.cell(row=row_idx, column=1, value=key).font = header_style
                worksheet.cell(row=row_idx, column=2, value=value).alignment = cell_alignment
            row_idx += 1
        # Saltar una fila en blanco
        row_idx += 1
        
        # Escribir sección de "CANON MENSUAL"
        worksheet.cell(row=row_idx, column=1, value="CANON MENSUAL").font = header_style
        row_idx += 1    
        for canon in self.canones_mensuales:
            worksheet.cell(row=row_idx, column=1, value="CANON MES")
            worksheet.cell(row=row_idx, column=2, value=canon["CANON MES"])
            worksheet.cell(row=row_idx, column=3, value=canon["Cantidad"])
            row_idx += 1
        
        # Saltar una fila en blanco
        
        # Escribir sección de "TOTAL CONTRATO"
        worksheet.cell(row=row_idx, column=1, value="TOTAL CONTRATO").font = header_style
        worksheet.cell(row=row_idx, column=3, value=self.total_contrato).border = border_top
        row_idx += 1
        
        # Saltar una fila en blanco
        row_idx += 1
        
        # Escribir sección de "DEPOSITOS EFECTUADOS"
        worksheet.cell(row=row_idx, column=1, value="DEPOSITOS EFECTUADOS").font = header_style
        row_idx += 1
        worksheet.cell(row=row_idx, column=1, value="Fecha").font = header_style
        worksheet.cell(row=row_idx, column=2, value="MODALIDAD").font = header_style
        worksheet.cell(row=row_idx, column=3, value="Cantidad").font = header_style
        row_idx += 1
        
        for deposito in self.depositos_efectuados:
            worksheet.cell(row=row_idx, column=1, value=deposito["Fecha"])
            worksheet.cell(row=row_idx, column=2, value=deposito["MODALIDAD"])
            worksheet.cell(row=row_idx, column=3, value=deposito["Cantidad"])
            row_idx += 1
        
        # Saltar una fila en blanco
        row_idx += 1
        
        # Escribir sección de "TOTAL DEPOSITADO"
        worksheet.cell(row=row_idx, column=1, value="TOTAL DEPOSITADO").font = header_style
        worksheet.cell(row=row_idx, column=3, value=self.total_depositado).border=border_top
        row_idx += 1
        
        # Saltar una fila en blanco
        row_idx += 1
        
        
        # Ajustar anchos de columna
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Obtener la letra de la columna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width
        
        # Alinear texto al centro para todas las celdas
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = cell_alignment
        
        # Aplicar formato de moneda a la columna de "Cantidad"
        for row in worksheet.iter_rows(min_row=4, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = currency_format
        
        # Guardar el archivo Excel
        if not os.path.exists(output_path):
            os.makedirs(output_path)
        
        # Guardar el archivo Excel en la ruta especificada
        fecha_actual = datetime.datetime.now().strftime("%Y-%m-%d")
        nombre_archivo = f"PagosContrato-{self.datos_inquilino['INQUILINO']}-{fecha_actual}.xlsx"
        ruta_completa = os.path.join(output_path, nombre_archivo)
        workbook.save(ruta_completa)
        
        print(f"Archivo '{nombre_archivo}' guardado en '{output_path}' exitosamente.")
        return {"nombre_archivo": nombre_archivo, "output_path": ruta_completa}


# Ejemplo de uso
if __name__ == "__main__":
    contrato = ContratoInquilinoExcel("contrato_inquilino.xlsx")
    
    # Configurar datos del contrato
    datos_inquilino = {
        "INQUILINO": "MARCOS LATORRACA",
        "N° CONTACTO": "0412-1927982",
        "CORREO": "",
        "INMUEBLE": "COCO GUAICA - TORRE A P122",
        "FECHA DE CONTRATO": "15/02/2022 AL 14/08/2022",
        "CANON": "USD 170,00",
        "DEPOSITO EN GARANTIA": "USD 450,00"
    }
    contrato.set_datos_inquilino(datos_inquilino)
    
    canones_mensuales = [
        {"CANON MES": "FEBRERO/MARZO 2022", "Cantidad": "170,00"},
        {"CANON MES": "MARZO/ABRIL 2022", "Cantidad": "170,00"},
        {"CANON MES": "ABRIL/MAYO 2022", "Cantidad": "170,00"},
        {"CANON MES": "MAYO/JUNIO 2022", "Cantidad": "170,00"},
        {"CANON MES": "JUNIO/JULIO 2022", "Cantidad": "170,00"},
        {"CANON MES": "JULIO/AGOSTO 2022", "Cantidad": "170,00"}
    ]
    contrato.set_canones_mensuales(canones_mensuales)
    
    total_contrato = "USD 1.020,00"
    contrato.set_total_contrato(total_contrato)
    
    depositos_efectuados = [
        {"Fecha": "2/17/2022", "MODALIDAD": "EFECTIVO", "Cantidad": "USD 170,00"},
        {"Fecha": "21/03/2022", "MODALIDAD": "EFECTIVO", "Cantidad": "USD 170,00"},
        {"Fecha": "21/04/2022", "MODALIDAD": "EFECTIVO", "Cantidad": "USD 170,00"},
        {"Fecha": "18/05/2022", "MODALIDAD": "EFECTIVO", "Cantidad": "USD 170,00"},
        {"Fecha": "21/06/2022", "MODALIDAD": "EFECTIVO", "Cantidad": "USD 170,00"},
        {"Fecha": "21/07/2022", "MODALIDAD": "EFECTIVO", "Cantidad": "USD 170,00"}
    ]
    contrato.set_depositos_efectuados(depositos_efectuados)
    
    total_depositado = "USD 1.020,00"
    contrato.set_total_depositado(total_depositado)
    
    # Generar el archivo Excel
    contrato.write_to_excel()
